import datetime
import os, re, pandas as pd
import time
import requests.exceptions
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import Authentication
import Requests
import config_path_file

reportDone = False
def findPhoneNumbersFile():
    """
    find file with phone numbers by regex
    :return: filename
    """
    datapath = config_path_file.DATA_PATH

    filename_regex = re.compile('bill_phone_numbers_' + '\d\d_\d\d_20\d\d' + '\.xls')

    for root, dirs, files in os.walk(datapath):
        for file in files:
            if filename_regex.match(file):
                return file

def createDailyReport(file, nds, output_folder, report_date=None):
    """
    create report with phone number and day outcome and save it into Excel file
    :return: dataframe with data
    """
    global reportDone
    if file is None:
        return
    print(file)
    print('nds=', nds)
    df = pd.read_excel(file)
    list = []
    print('start creating report')
    print('reportDone=', reportDone)
    phones = df.get('Абонентский номер')
    print(phones)
    if reportDone:
        return
    if report_date == None:
        report_date = datetime.datetime.now()
    for phone in phones:
        i = 0
        successful = False
        while not successful:
            print(successful)
            try:

                createdRequestDailyExpenses = Requests.create_request(flag='BILLS_BY_MSISDN',
                                                                     params_list=[phone, report_date - datetime.timedelta(1), report_date])

                next_month = report_date.replace(day=28) + datetime.timedelta(days=4)
                createdRequestMonthlyExpenses = Requests.create_request(flag='BILLS_BY_MSISDN',
                                                                       params_list=[phone, report_date.replace(day=1), report_date])
                print('STATUS_CODES', createdRequestDailyExpenses.status_code, createdRequestMonthlyExpenses.status_code)
                if createdRequestDailyExpenses.status_code == requests.codes.ok and createdRequestMonthlyExpenses.status_code == requests.codes.ok:
                    successful = True
                    month_amount = summarize(createdRequestMonthlyExpenses)
                    amount = summarize(createdRequestDailyExpenses)
                    index = int(df[df['Абонентский номер']==phone].index.values.astype(int)[0])
                    name = df.get('ФИО')[index]
                    commentary = df.get('Комментарий')[index]
                    limit = df.get('Лимит')[index]
                    nds_percentage = (1-nds/100)
                    no_nds_value = amount* nds_percentage
                    no_nds_month_value = month_amount * nds_percentage
                    limit_excess = no_nds_month_value - limit
                    if limit_excess < 0:
                        limit_excess = 0
                    data = [phone, name, commentary, amount, no_nds_value, month_amount, no_nds_month_value, limit_excess]
                    print(data)
                    list.append(data)
                else:
                    time.sleep(10)
            except requests.exceptions.HTTPError as e:
                print(e.args)
                if (int(e.args[0][0:3]) == 429):
                    print('sleep')
                    time.sleep(10)
                    continue
                elif (int(e.args[0][0:3]) == 401) and i < 2:
                    Authentication.loginUser(Authentication._login, Authentication._password)
                    time.sleep(60)
                    i+=1
                    continue
                else:
                    print('wtf')
                    successful = True
                    continue


    path = uniquify(output_folder + '/report_' + report_date.date().strftime("%d_%m_%y") + '.xlsx')
    new_df = pd.DataFrame(list, columns = ['Абонентский номер', 'ФИО', 'Комментарий', 'с НДС', 'без НДС','с НДС',  "без НДС",  'Превышение лимита'])
    print(new_df)
    if not new_df.empty:
        print('not empty')
        wb = openpyxl.Workbook()
        ws = wb.active

        for row in dataframe_to_rows(new_df, index=False, header=True):
            ws.append(row)

        month = ''
        match report_date.month:
            case 1:
                month = "Январь"
            case 2:
                month = "Февраль"
            case 3:
                month = "Март"
            case 4:
                month = "Апрель"
            case 5:
                month = "Май"
            case 6:
                month = "Июнь"
            case 7:
                month = "Июль"
            case 8:
                month = "Август"
            case 9:
                month = "Сентябрь"
            case 10:
                month = "Октябрь"
            case 11:
                month = "Ноябрь"
            case 12:
                month = "Декабрь"

        print(month)

        ws.insert_rows(1, amount=2)
        ws.merge_cells('D1:G1')
        ws.cell(1, 4).value = f'Расходы, {month}'
        ws.merge_cells('D2:E2')
        ws.cell(2, 4).value = 'За день'
        ws.merge_cells('F2:G2')
        ws.cell(2, 6).value = 'С начала месяца'


        i = 0
        for column in ws.columns:
            i+=1
            max_length = 0
            col = openpyxl.utils.cell.get_column_letter(i)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            width = max_length + 2
            ws.column_dimensions[col].width = width

        print(ws)
        os.chmod(path, 755)
        wb.save(path)
    print('done')
    reportDone = True
    return new_df

def uniquify(path):
    filename, extension = os.path.splitext(path)
    counter = 1
    while os.path.exists(path):
        path = filename + " (" + str(counter) + ")" + extension
        counter += 1
    return path

def summarize(rqst):
    sum = 0
    for oper in rqst.json()['Usages']:
        if oper['type'] != 'income':
            sum += oper['amount']
    return sum

